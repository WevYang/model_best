#!/usr/bin/env python3
"""
Hallucination benchmark analysis pipeline.
Reads all model PDFs/HTMLs, extracts benchmark mentions & scores,
generates matrix, heatmap, Excel tables, taxonomy, and enhanced manifest.
"""
import json, re, os, sys, hashlib, datetime, warnings, io
from pathlib import Path
from collections import defaultdict

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment,
                              Border, Side, GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference

warnings.filterwarnings("ignore")

BASE    = Path(__file__).parent
OUT     = BASE / "analysis_results"
OUT.mkdir(exist_ok=True)

# ── 1. CONSTANTS ──────────────────────────────────────────────────────────────

# Hallucination / factuality benchmark definitions
BENCHMARKS = {
    "CHAIR":          {"alias": ["chair","object hallucination in image"],
                       "type": "object_hal", "metric": "CHAIR_i/CHAIR_s"},
    "POPE":           {"alias": ["pope","polling-based object probing"],
                       "type": "object_hal", "metric": "Accuracy/F1"},
    "HallusionBench": {"alias": ["hallusionbench","hallusion bench","hallusion"],
                       "type": "visual_hal", "metric": "Accuracy"},
    "MMHal-Bench":    {"alias": ["mmhal","mmhal-bench","mmhal bench"],
                       "type": "multimodal_hal", "metric": "Score(0-6)"},
    "AMBER":          {"alias": ["amber","an llm-free multi-dimensional benchmark"],
                       "type": "comprehensive_hal", "metric": "AMBER Score"},
    "FaithScore":     {"alias": ["faithscore","faith score","faithfulness score"],
                       "type": "faithfulness", "metric": "FaithScore"},
    "OCRBench":       {"alias": ["ocrbench","ocr bench"],
                       "type": "ocr_hal", "metric": "OCRBench Score"},
    "OCRBench-v2":    {"alias": ["ocrbench-v2","ocrbench v2","ocrbench2"],
                       "type": "ocr_hal", "metric": "OCRBench-v2 Score"},
    "FACTS-Grounding":{"alias": ["facts grounding","facts-grounding","factuality grounding"],
                       "type": "grounding_hal", "metric": "Factuality Score"},
    "CharXiv":        {"alias": ["charxiv","chart ","chartqa"],
                       "type": "chart_hal", "metric": "Accuracy"},
    "Video-MME":      {"alias": ["video-mme","videomme","video mme"],
                       "type": "video_hal", "metric": "Accuracy"},
    "LongVideoBench": {"alias": ["longvideobench","long video bench"],
                       "type": "video_hal", "metric": "Accuracy"},
    "TempCompass":    {"alias": ["tempcompass","temporal compass"],
                       "type": "video_temporal_hal", "metric": "Accuracy"},
    "SimpleVQA":      {"alias": ["simplevqa","simple vqa","simple-vqa"],
                       "type": "visual_hal", "metric": "Accuracy"},
    "ScienceQA":      {"alias": ["scienceqa","science qa"],
                       "type": "reasoning_hal", "metric": "Accuracy"},
    "MMBench":        {"alias": ["mmbench","mm-bench"],
                       "type": "general_mm", "metric": "Accuracy"},
    "MMStar":         {"alias": ["mmstar","mm-star"],
                       "type": "general_mm", "metric": "Accuracy"},
    "MathVista":      {"alias": ["mathvista","math vista"],
                       "type": "reasoning_hal", "metric": "Accuracy"},
    "AI2D":           {"alias": ["ai2d"],
                       "type": "reasoning_hal", "metric": "Accuracy"},
    "TextVQA":        {"alias": ["textvqa","text vqa"],
                       "type": "ocr_hal", "metric": "Accuracy"},
    "DocVQA":         {"alias": ["docvqa","doc vqa"],
                       "type": "ocr_hal", "metric": "Accuracy"},
    "ChartQA":        {"alias": ["chartqa","chart qa"],
                       "type": "chart_hal", "metric": "Accuracy"},
    "InfoVQA":        {"alias": ["infovqa","info vqa"],
                       "type": "ocr_hal", "metric": "Accuracy"},
    "VQAv2":          {"alias": ["vqav2","vqa v2","vqa-v2"],
                       "type": "general_mm", "metric": "Accuracy"},
    "NoCaps":         {"alias": ["nocaps","no caps"],
                       "type": "captioning_hal", "metric": "CIDEr"},
    "COCO-Cap":       {"alias": ["coco caption","coco cap","ms-coco"],
                       "type": "captioning_hal", "metric": "CIDEr"},
    "RefCOCO":        {"alias": ["refcoco","ref coco"],
                       "type": "grounding_hal", "metric": "Accuracy"},
    "MMVet":          {"alias": ["mm-vet","mmvet"],
                       "type": "comprehensive_hal", "metric": "Score"},
    "LLaVA-Bench":    {"alias": ["llava-bench","llava bench"],
                       "type": "comprehensive_hal", "metric": "Score"},
    "MMMU":           {"alias": ["mmmu"],
                       "type": "reasoning_hal", "metric": "Accuracy"},
    "Winoground":     {"alias": ["winoground","wino ground"],
                       "type": "relation_hal", "metric": "Accuracy"},
    "ARO":            {"alias": ["aro ","attribution,relation,order"],
                       "type": "relation_hal", "metric": "Accuracy"},
    "VisWiz":         {"alias": ["viswiz","vis-wiz"],
                       "type": "general_mm", "metric": "Accuracy"},
}

# Hallucination TYPE taxonomy
HAL_TYPES = ["object", "attribute", "relation", "OCR/text", "grounding",
             "reasoning/math", "video/temporal", "chart/document",
             "captioning", "general_mm"]

# Benchmark → hallucination type mapping
BM_TO_HAL_TYPES = {
    "CHAIR":           ["object"],
    "POPE":            ["object"],
    "HallusionBench":  ["object","attribute","relation"],
    "MMHal-Bench":     ["object","attribute","relation"],
    "AMBER":           ["object","attribute","relation"],
    "FaithScore":      ["attribute","relation"],
    "OCRBench":        ["OCR/text"],
    "OCRBench-v2":     ["OCR/text"],
    "FACTS-Grounding": ["grounding"],
    "CharXiv":         ["chart/document","reasoning/math"],
    "Video-MME":       ["video/temporal","object"],
    "LongVideoBench":  ["video/temporal"],
    "TempCompass":     ["video/temporal"],
    "SimpleVQA":       ["object","attribute"],
    "TextVQA":         ["OCR/text"],
    "DocVQA":          ["OCR/text","chart/document"],
    "ChartQA":         ["chart/document"],
    "InfoVQA":         ["OCR/text"],
    "NoCaps":          ["captioning","object"],
    "COCO-Cap":        ["captioning","object"],
    "RefCOCO":         ["grounding"],
    "MMVet":           ["object","attribute","relation","reasoning/math"],
    "MMBench":         ["general_mm"],
    "MMStar":          ["general_mm"],
    "MathVista":       ["reasoning/math"],
    "ScienceQA":       ["reasoning/math"],
    "AI2D":            ["reasoning/math","chart/document"],
    "VQAv2":           ["general_mm"],
    "MMMU":            ["reasoning/math"],
    "LLaVA-Bench":     ["object","attribute","relation"],
    "Winoground":      ["relation"],
    "ARO":             ["relation"],
    "VisWiz":          ["general_mm"],
}

# ── 2. PDF / HTML TEXT EXTRACTION ─────────────────────────────────────────────
def extract_pdf_text(path: Path, max_pages=40) -> str:
    """Extract PDF text via pdftotext (fast) with pdfminer fallback."""
    import subprocess
    try:
        result = subprocess.run(
            ["pdftotext", "-l", str(max_pages), str(path), "-"],
            capture_output=True, timeout=20, text=True, errors="replace"
        )
        if result.returncode == 0 and result.stdout.strip():
            return result.stdout
    except Exception:
        pass
    # pdfminer fallback (slower)
    try:
        from pdfminer.high_level import extract_text
        import signal
        def _h(s, f): raise TimeoutError()
        signal.signal(signal.SIGALRM, _h)
        signal.alarm(15)
        try:
            text = extract_text(str(path), maxpages=20) or ""
        finally:
            signal.alarm(0)
        return text
    except Exception as e:
        return f"[ERROR: {e}]"

def extract_html_text(path: Path) -> str:
    """Strip HTML tags and return plain text."""
    try:
        content = path.read_text(errors="replace")
        # Remove script/style
        content = re.sub(r"<(script|style)[^>]*>.*?</\1>", " ", content, flags=re.S|re.I)
        content = re.sub(r"<[^>]+>", " ", content)
        content = re.sub(r"&[a-zA-Z]+;", " ", content)
        content = re.sub(r"\s+", " ", content)
        return content[:200000]
    except Exception as e:
        return f"[ERROR: {e}]"

def extract_text(path: Path) -> str:
    if not path or not path.exists():
        return ""
    if path.suffix.lower() == ".pdf":
        return extract_pdf_text(path)
    elif path.suffix.lower() in (".html", ".htm", ".md"):
        return extract_html_text(path)
    return path.read_text(errors="replace")[:200000]

# ── 3. SCORE EXTRACTION ───────────────────────────────────────────────────────
def extract_scores_for_benchmark(text: str, bm_name: str) -> list[tuple[str, float, str]]:
    """
    Return list of (context, value, raw_match) for a benchmark in text.
    Uses regex patterns around benchmark name mentions.
    """
    aliases = BENCHMARKS.get(bm_name, {}).get("alias", [bm_name.lower()])
    results = []
    # Pattern: benchmark name followed (within ~100 chars) by a number
    for alias in aliases:
        escaped = re.escape(alias)
        # Forward: "POPE ... 86.7"
        fwd = re.findall(
            rf"(?i){escaped}.{{0,120}}?(\d{{1,3}}(?:\.\d{{1,2}})?)\b",
            text)
        # Backward: "86.7 ... POPE"
        bwd = re.findall(
            rf"(?i)(\d{{1,3}}(?:\.\d{{1,2}})?).{{0,120}}?{escaped}",
            text)
        for v in fwd + bwd:
            try:
                val = float(v)
                if 0 < val <= 100:
                    results.append((alias, val, v))
            except Exception:
                pass
    # Deduplicate by value
    seen = set()
    out = []
    for item in results:
        k = round(item[1], 1)
        if k not in seen:
            seen.add(k)
            out.append(item)
    return out[:5]  # top 5 distinct

def find_benchmark_mentions(text: str, bm_name: str) -> bool:
    aliases = BENCHMARKS.get(bm_name, {}).get("alias", [bm_name.lower()])
    text_lo = text.lower()
    return any(a.lower() in text_lo for a in aliases)

# ── 4. LOAD MANIFEST & EXTRACT ALL TEXT ───────────────────────────────────────
print("Loading manifest…")
manifest = json.loads((BASE / "_download_manifest.json").read_text())

# Only model entries (not benchmark papers)
MODEL_ENTRIES = [e for e in manifest
                 if e["doc_type"] not in ("benchmark_paper",)
                 and "image_generation" not in e.get("notes","").lower()]

# Build model name → entry mapping (deduplicate, prefer PDF)
model_map: dict[str, dict] = {}
for e in MODEL_ENTRIES:
    name = e["model_or_benchmark_name"]
    if name not in model_map:
        model_map[name] = e
    else:
        # prefer PDF over HTML
        existing = model_map[name]
        if e["local_path"].endswith(".pdf") and not existing["local_path"].endswith(".pdf"):
            model_map[name] = e

# Canonical model order (roughly chronological)
CANONICAL_ORDER = [
    # OpenAI
    "GPT-4","GPT-4V","GPT-4o","GPT-4.1","GPT-5","GPT-5.2","GPT-5.5","GPT-5.5-Instant",
    # Anthropic
    "Claude-3.5","Claude-3.5-Haiku-Sonnet","Claude-4","Claude-4.1",
    "Claude-Sonnet-4.6","Claude-Opus-4.6","Claude-Opus-4.7","Claude-Mythos-Preview",
    # Google
    "Gemini-1.5","Gemini-2.5","Gemini-2.0-Flash",
    "Gemini-2.5-Flash","Gemini-2.5-Pro",
    "Gemini-3-Pro","Gemini-3.1-Pro","Gemini-3.1-Flash-Lite",
    "Gemini-3.1-Flash-Audio","Gemini-3.1-Flash-Image","Gemini-3.5-Flash","Gemini-Omni-Flash",
    "Gemma 3","Gemma-3n","PaliGemma","PaliGemma 2",
    # Meta
    "LLaVA","LLaVA-OneVision","LLaVA-OneVision-2","Llama-3.2-Vision","Llama-4",
    # Alibaba / Qwen
    "Qwen2.5-VL","Qwen3-VL","Qwen2.5-Omni","Qwen3-Omni","Qwen3.5-Omni","Qwen3.6","Qwen3.7",
    # DeepSeek
    "DeepSeek-VL","DeepSeek-VL2",
    # NVIDIA / Microsoft
    "NVLM 1.0","Phi-4-multimodal","Kosmos-1","Kosmos-2",
    # Shanghai AI Lab
    "InternVL","InternVL-1.5","InternVL-2.5","InternVL-3",
    # ModelBest / Tsinghua
    "MiniCPM","MiniCPM-Llama3-V","MiniCPM-V","MiniCPM-o",
    # ByteDance
    "SEED-LLaMA","VILA","VILA-1.5",
    # Zhipu
    "CogVLM","GLM-4.5V","GLM-4.1V-Thinking","GLM-4.6V",
    # Misc
    "BLIP-2","Bunny","Emu","Emu2","Idefics2",
    "Kimi-VL","Molmo","MiMo-VL","Ovis2.5","Ovis-U1",
    "PaLI","PaLI-X","PaLM-E","Pixtral-12B",
    # OCR
    "HunyuanOCR","DeepSeek-OCR","PaddleOCR-VL","MinerU2.5",
]

# Get ordered list of models present in manifest
all_model_names = list(model_map.keys())
ordered_models = [m for m in CANONICAL_ORDER if m in model_map]
remaining = [m for m in all_model_names if m not in ordered_models]
ordered_models += sorted(remaining)

print(f"Total models to analyze: {len(ordered_models)}")
print(f"Total benchmarks to check: {len(BENCHMARKS)}")

# Extract text for all models (with progress)
print("\nExtracting text from documents…")
model_texts: dict[str, str] = {}
for name in ordered_models:
    entry = model_map[name]
    lp = entry.get("local_path", "")
    if not lp:
        model_texts[name] = ""
        print(f"  [no file] {name}")
        continue
    path = BASE.parent / lp
    text = extract_text(path)
    model_texts[name] = text
    status = "PDF" if lp.endswith(".pdf") else "HTML"
    print(f"  [{status} {len(text)//1000}k chars] {name}")

# ── 5. BUILD MODEL × BENCHMARK MATRIX ─────────────────────────────────────────
print("\nBuilding model × benchmark matrix…")

# Primary hallucination benchmarks for the main matrix
PRIMARY_BENCHMARKS = [
    "CHAIR","POPE","HallusionBench","MMHal-Bench","AMBER","FaithScore",
    "OCRBench","OCRBench-v2","FACTS-Grounding","CharXiv",
    "Video-MME","LongVideoBench","TempCompass","SimpleVQA",
    # Common capability benchmarks (proxy)
    "TextVQA","DocVQA","ChartQA","InfoVQA",
    "MMBench","MMStar","MathVista","MMMU","MMVet",
    "NoCaps","COCO-Cap","VQAv2","ScienceQA","AI2D",
]

matrix_rows = []
score_data: dict[str, dict[str, list]] = defaultdict(lambda: defaultdict(list))

for model_name in ordered_models:
    text = model_texts[model_name]
    entry = model_map[model_name]
    row = {"Model": model_name,
           "Organization": entry.get("organization",""),
           "doc_type": entry.get("doc_type",""),
           "modality": entry.get("modality",""),
           "source_type": entry.get("source_type",""),
           "local_path": entry.get("local_path",""),
           "notes": entry.get("notes",""),
    }
    for bm in PRIMARY_BENCHMARKS:
        mentioned = find_benchmark_mentions(text, bm)
        scores = extract_scores_for_benchmark(text, bm) if mentioned else []

        # Scoring: 2=explicitly evaluated, 1=mentioned/proxy, 0=not mentioned
        if scores:
            val = 2
        elif mentioned:
            val = 1
        else:
            val = 0
        row[bm] = val

        # Store best score value
        if scores:
            best_score = max(s[1] for s in scores)
            score_data[model_name][bm] = best_score

    matrix_rows.append(row)

matrix_df = pd.DataFrame(matrix_rows)
matrix_df.set_index("Model", inplace=True)

# ── 6. HALLUCINATION TYPE COVERAGE ────────────────────────────────────────────
print("Computing hallucination type coverage…")

hal_type_rows = []
for model_name in ordered_models:
    text = model_texts[model_name]
    entry = model_map[model_name]
    type_coverage = defaultdict(int)

    for bm, bm_info in BENCHMARKS.items():
        if find_benchmark_mentions(text, bm):
            hal_types = BM_TO_HAL_TYPES.get(bm, [])
            scores = extract_scores_for_benchmark(text, bm)
            weight = 2 if scores else 1
            for ht in hal_types:
                type_coverage[ht] = max(type_coverage[ht], weight)

    row = {"Model": model_name}
    for ht in HAL_TYPES:
        row[ht] = type_coverage.get(ht, 0)
    hal_type_rows.append(row)

haltype_df = pd.DataFrame(hal_type_rows).set_index("Model")

# ── 7. OFFICIAL EXPERIMENT RESULTS TABLE ──────────────────────────────────────
print("Extracting official experiment results…")

# Enhanced extraction: try to find actual score values with context
SCORE_PATTERNS = {
    "POPE": [
        r"pope[^\d]{0,30}(\d{2,3}(?:\.\d{1,2})?)",
        r"(\d{2,3}(?:\.\d{1,2})?)[^\d]{0,30}pope",
    ],
    "CHAIR": [
        r"chair[_\s-]?[is][^\d]{0,20}(\d{1,2}(?:\.\d{1,2})?)",
        r"(\d{1,2}(?:\.\d{1,2})?)[^\d]{0,20}chair[_\s-]?[is]",
    ],
    "HallusionBench": [
        r"hallusionbench[^\d]{0,40}(\d{2,3}(?:\.\d{1,2})?)",
        r"hallusion[^\d]{0,40}(\d{2,3}(?:\.\d{1,2})?)",
    ],
    "OCRBench": [
        r"ocrbench[^\d]{0,30}(\d{2,3}(?:\.\d{1,2})?)",
        r"(\d{3})[^\d]{0,30}ocrbench",
    ],
    "TextVQA": [
        r"textvqa[^\d]{0,30}(\d{2,3}(?:\.\d{1,2})?)",
        r"text.?vqa[^\d]{0,30}(\d{2,3}(?:\.\d{1,2})?)",
    ],
    "DocVQA": [
        r"docvqa[^\d]{0,30}(\d{2,3}(?:\.\d{1,2})?)",
        r"doc.?vqa[^\d]{0,30}(\d{2,3}(?:\.\d{1,2})?)",
    ],
    "ChartQA": [
        r"chartqa[^\d]{0,30}(\d{2,3}(?:\.\d{1,2})?)",
        r"chart.?qa[^\d]{0,30}(\d{2,3}(?:\.\d{1,2})?)",
    ],
    "MMBench": [
        r"mmbench[^\d]{0,30}(\d{2,3}(?:\.\d{1,2})?)",
    ],
    "MMVet": [
        r"mm.?vet[^\d]{0,30}(\d{2,3}(?:\.\d{1,2})?)",
    ],
    "MathVista": [
        r"mathvista[^\d]{0,30}(\d{2,3}(?:\.\d{1,2})?)",
        r"math.?vista[^\d]{0,30}(\d{2,3}(?:\.\d{1,2})?)",
    ],
    "MMMU": [
        r"\bmmmu\b[^\d]{0,30}(\d{2,3}(?:\.\d{1,2})?)",
    ],
    "Video-MME": [
        r"video.?mme[^\d]{0,30}(\d{2,3}(?:\.\d{1,2})?)",
    ],
    "NoCaps": [
        r"nocaps[^\d]{0,30}(\d{2,3}(?:\.\d{1,2})?)",
    ],
    "VQAv2": [
        r"vqa.?v2[^\d]{0,30}(\d{2,3}(?:\.\d{1,2})?)",
    ],
    "AI2D": [
        r"\bai2d\b[^\d]{0,30}(\d{2,3}(?:\.\d{1,2})?)",
    ],
    "ScienceQA": [
        r"scienceqa[^\d]{0,30}(\d{2,3}(?:\.\d{1,2})?)",
        r"science.?qa[^\d]{0,30}(\d{2,3}(?:\.\d{1,2})?)",
    ],
    "MMStar": [
        r"mmstar[^\d]{0,30}(\d{2,3}(?:\.\d{1,2})?)",
        r"mm.?star[^\d]{0,30}(\d{2,3}(?:\.\d{1,2})?)",
    ],
}

results_rows = []
for model_name in ordered_models:
    text_lo = model_texts[model_name].lower()
    entry = model_map[model_name]

    row = {
        "Model": model_name,
        "Organization": entry.get("organization", ""),
        "doc_type": entry.get("doc_type", ""),
        "source_file": Path(entry.get("local_path","")).name,
        "modality": entry.get("modality",""),
    }

    for bm, patterns in SCORE_PATTERNS.items():
        best = None
        for pat in patterns:
            matches = re.findall(pat, text_lo)
            for m in matches:
                try:
                    v = float(m)
                    if 0 < v <= 100:
                        if best is None or abs(v - 50) < abs(best - 50):
                            best = v
                except Exception:
                    pass
        row[bm] = best

    results_rows.append(row)

results_df = pd.DataFrame(results_rows).set_index("Model")

# ── 8. BENCHMARK TAXONOMY ─────────────────────────────────────────────────────
print("Building benchmark taxonomy…")

TAXONOMY = {
    "CHAIR": {
        "full_name": "Caption Hallucination Assessment with Image Relevance",
        "arxiv": "1809.02157",
        "year": 2018,
        "type": "Object Hallucination",
        "task": "Image captioning evaluation — counts hallucinated objects",
        "metrics": "CHAIR_i (per-instance), CHAIR_s (per-sentence) ↓ lower=better",
        "scale": "~5K COCO images; calculates % of hallucinated object words",
        "open_source": "Yes (eval code open)",
        "strengths": "Simple, interpretable, object-level precision; widely cited",
        "weaknesses": "Only objects (no attributes/relations); COCO-biased; not scalable to diverse caps",
        "common_pitfalls": "Using non-COCO images; confusion between CHAIR_i and CHAIR_s; outdated object list",
        "hal_types": "object",
    },
    "POPE": {
        "full_name": "Polling-based Object Probing Evaluation",
        "arxiv": "2305.10355",
        "year": 2023,
        "type": "Object Hallucination",
        "task": "Yes/No binary QA: 'Is there a [object] in the image?'",
        "metrics": "Accuracy, Precision, Recall, F1 (% correct Y/N answers)",
        "scale": "~9K questions across 3 sampling strategies (random/popular/adversarial)",
        "open_source": "Yes (DAMO Academy GitHub)",
        "strengths": "Easy to evaluate; adversarial set exposes frequency bias; widely adopted",
        "weaknesses": "Binary only; misses attribute/relation errors; model can game with prior",
        "common_pitfalls": "Reporting only random split; not separating adversarial F1",
        "hal_types": "object",
    },
    "HallusionBench": {
        "full_name": "HallusionBench: Advanced Diagnostic Suite for Hallucination",
        "arxiv": "2310.14566",
        "year": 2023,
        "type": "Visual Hallucination (multi-type)",
        "task": "Paired questions: same question, different visual context (augmented/original)",
        "metrics": "Question Accuracy, Figure Accuracy, Overall Score",
        "scale": "~1.1K visual Q&A pairs across 346 images",
        "open_source": "Yes (GitHub)",
        "strengths": "Adversarially designed; covers visual language confusion; language-only test pairs",
        "weaknesses": "Small scale; requires manual verification of augmented images",
        "common_pitfalls": "Mixing language-only and visual subtasks in single number",
        "hal_types": "object, attribute, relation",
    },
    "MMHal-Bench": {
        "full_name": "Multimodal Hallucination Benchmark (LLaVA-RLHF paper)",
        "arxiv": "2309.14525",
        "year": 2023,
        "type": "Open-ended Multimodal Hallucination",
        "task": "Open-ended VQA; GPT-4 rates hallucination on 0-6 scale",
        "metrics": "Score (0-6 by GPT-4); Hallucination Rate",
        "scale": "96 question-image pairs across 8 question types",
        "open_source": "Yes (ByteDance GitHub)",
        "strengths": "Open-ended; captures nuanced hallucinations; GPT-4 judged",
        "weaknesses": "Very small scale (96 samples); GPT-4 judging cost; judge variance",
        "common_pitfalls": "Using GPT-3.5 instead of GPT-4 as judge; scale too small for significance",
        "hal_types": "object, attribute, relation",
    },
    "AMBER": {
        "full_name": "An LLM-Free Multi-Dimensional Benchmark for Hallucination Evaluation",
        "arxiv": "2311.07397",
        "year": 2023,
        "type": "Comprehensive Hallucination",
        "task": "Generative (free-form description) + Discriminative (Y/N) tasks",
        "metrics": "AMBER Score (composite), Accuracy (discriminative), Hallucination Rate (generative)",
        "scale": "~1K images, generative + discriminative tasks",
        "open_source": "Yes (Beihang University GitHub)",
        "strengths": "LLM-free evaluation (no API cost); multi-dimensional; both gen+discrim",
        "weaknesses": "Less nuanced than GPT-4-judged; newer so less baseline coverage",
        "common_pitfalls": "Reporting only discriminative part; confusing generative metric direction",
        "hal_types": "object, attribute, relation",
    },
    "FaithScore": {
        "full_name": "FaithScore: Fine-grained Faithfulness Evaluation",
        "arxiv": "2311.01477",
        "year": 2023,
        "type": "Faithfulness / Attribute Hallucination",
        "task": "Atomic sub-sentence factuality checking against image content",
        "metrics": "FaithScore (precision of atomic facts; 0-1)",
        "scale": "ActivityNet Captions + others; custom evaluation pipeline",
        "open_source": "Partial (evaluation code)",
        "strengths": "Fine-grained atomic fact checking; captures attribute errors missed by CHAIR",
        "weaknesses": "Requires NLP pipeline; slower evaluation; less widely adopted",
        "common_pitfalls": "Conflating with CHAIR; dependency on NLP parser quality",
        "hal_types": "attribute, relation",
    },
    "OCRBench": {
        "full_name": "OCRBench: On the Hidden Mystery of OCR in Large Multimodal Models",
        "arxiv": "2305.07895",
        "year": 2023,
        "type": "OCR / Text Recognition Hallucination",
        "task": "31 OCR tasks: text recognition, formula, table, scene text, etc.",
        "metrics": "OCRBench Score (0-1000), per-task accuracy",
        "scale": "~29K samples across 31 subtasks",
        "open_source": "Yes (Shanghai AI Lab GitHub)",
        "strengths": "Comprehensive OCR coverage; multiple subtasks; widely reported",
        "weaknesses": "Score range (0-1000) non-intuitive; doesn't cover all document types",
        "common_pitfalls": "Comparing score without normalizing to 100; ignoring subtask breakdown",
        "hal_types": "OCR/text",
    },
    "OCRBench-v2": {
        "full_name": "OCRBench-v2: Comprehensive OCR Evaluation Suite (Extended)",
        "arxiv": "2404.00773",
        "year": 2024,
        "type": "OCR / Document Understanding",
        "task": "Extended OCR tasks including layout, formula, handwriting",
        "metrics": "Score (0-1000), per-category breakdown",
        "scale": "~10K+ samples, more diverse document types",
        "open_source": "Yes",
        "strengths": "More comprehensive than v1; covers layout understanding",
        "weaknesses": "Very new; fewer baseline comparisons available",
        "common_pitfalls": "Conflating v1 and v2 scores (different scale/tasks)",
        "hal_types": "OCR/text, chart/document",
    },
    "FACTS-Grounding": {
        "full_name": "FACTS Grounding: A Benchmark for Evaluating Factual Grounding",
        "arxiv": "2406.19523",
        "year": 2024,
        "type": "Factual Grounding / Attribution Hallucination",
        "task": "Evaluate factual grounding of long-form answers against source documents",
        "metrics": "Factuality Score (0-1), AutoFAIR rating",
        "scale": "~860 examples across web search and document contexts",
        "open_source": "Yes (Google DeepMind)",
        "strengths": "Long-form factuality; grounded attribution testing; real-world relevance",
        "weaknesses": "Document-centric; less applicable to pure vision tasks",
        "common_pitfalls": "Applying to visual-only tasks; AutoFAIR vs manual rating discrepancy",
        "hal_types": "grounding",
    },
    "CharXiv": {
        "full_name": "CharXiv: Charting Territories in Chart Understanding",
        "arxiv": "2406.18521",
        "year": 2024,
        "type": "Chart Hallucination / Document Understanding",
        "task": "Descriptive and reasoning questions about real arXiv charts",
        "metrics": "Descriptive Accuracy, Reasoning Accuracy, Overall",
        "scale": "~2.3K questions from ~500 scientific charts",
        "open_source": "Yes (Princeton NLP)",
        "strengths": "Real-world scientific charts; tests reasoning not just pattern matching",
        "weaknesses": "Scientific domain bias; requires strong math/science knowledge",
        "common_pitfalls": "Treating descriptive and reasoning subtasks as interchangeable",
        "hal_types": "chart/document, reasoning/math",
    },
    "Video-MME": {
        "full_name": "Video-MME: First-Ever Comprehensive Video LMM Evaluation",
        "arxiv": "2405.21075",
        "year": 2024,
        "type": "Video Understanding / Temporal Hallucination",
        "task": "MCQ across short/medium/long videos with/without subtitles",
        "metrics": "Accuracy (w/ and w/o subtitles), per-duration breakdown",
        "scale": "~900 videos, ~2.7K QA pairs, 30 domains",
        "open_source": "Yes",
        "strengths": "Multi-duration; diverse domains; subtitle ablation; widely reported",
        "weaknesses": "MCQ format; video acquisition restrictions for some subsets",
        "common_pitfalls": "Reporting w/ subtitles only; not separating short/medium/long",
        "hal_types": "video/temporal, object",
    },
    "LongVideoBench": {
        "full_name": "LongVideoBench: Interleaved Video-Language Benchmark",
        "arxiv": "2407.15754",
        "year": 2024,
        "type": "Long Video / Temporal Reasoning",
        "task": "Multiple-choice referring reasoning on long videos (15s–60min)",
        "metrics": "Accuracy across duration bins",
        "scale": "~3.7K QA pairs from 1.5K videos",
        "open_source": "Yes",
        "strengths": "Very long videos; interleaved subtitles; temporal reasoning required",
        "weaknesses": "Evaluation infrastructure complex; video downloads may be restricted",
        "common_pitfalls": "Conflating with Video-MME; neglecting video-only vs w-subtitle gap",
        "hal_types": "video/temporal",
    },
    "TempCompass": {
        "full_name": "TempCompass: Do Video LLMs Really Understand Videos?",
        "arxiv": "2406.09436",
        "year": 2024,
        "type": "Temporal Comprehension Hallucination",
        "task": "Yes/No, Caption Matching, MCQ, Caption Generation on temporal video content",
        "metrics": "Accuracy per task type and temporal aspect",
        "scale": "~7.5K QA pairs from ~410 videos",
        "open_source": "Yes (Peking University)",
        "strengths": "Temporal-focused; multi-task format; exposes temporal reasoning failures",
        "weaknesses": "Small video set; task format diversity complicates aggregation",
        "common_pitfalls": "Averaging over task types with different baselines",
        "hal_types": "video/temporal",
    },
    "SimpleVQA": {
        "full_name": "SimpleVQA: Measuring Hallucination in LVLMs",
        "arxiv": "2310.02255",
        "year": 2023,
        "type": "Object / Visual Hallucination",
        "task": "Short-answer VQA focused on factual correctness",
        "metrics": "Accuracy, Hallucination Rate",
        "scale": "~3K QA pairs across common objects/attributes",
        "open_source": "Yes",
        "strengths": "Simple evaluation; focuses on factual errors; controllable",
        "weaknesses": "Less adversarial; may not capture subtle hallucinations",
        "common_pitfalls": "Using different answer normalization leading to incomparable scores",
        "hal_types": "object, attribute",
    },
}

taxonomy_rows = []
for bm_name, info in TAXONOMY.items():
    row = {"Benchmark": bm_name, **info}
    # Count how many models report this benchmark
    count = sum(1 for model in ordered_models
                if find_benchmark_mentions(model_texts[model], bm_name))
    row["#Models Reporting"] = count
    taxonomy_rows.append(row)

taxonomy_df = pd.DataFrame(taxonomy_rows).set_index("Benchmark")

# ── 9. ENHANCED MANIFEST ──────────────────────────────────────────────────────
print("Building enhanced manifest…")

enhanced = []
for entry in manifest:
    name = entry["model_or_benchmark_name"]
    lp = entry.get("local_path","")
    text = model_texts.get(name, "")

    # Find all benchmarks mentioned
    bm_mentioned = [bm for bm in BENCHMARKS if find_benchmark_mentions(text, bm)]

    # Collect score data
    bm_scores = {}
    for bm in bm_mentioned:
        scores = extract_scores_for_benchmark(text, bm)
        if scores:
            bm_scores[bm] = max(s[1] for s in scores)

    # Hallucination types covered
    hal_types_covered = set()
    for bm in bm_mentioned:
        hal_types_covered.update(BM_TO_HAL_TYPES.get(bm, []))

    enh = dict(entry)
    enh["benchmarks_mentioned"] = bm_mentioned
    enh["benchmark_scores_extracted"] = bm_scores
    enh["hallucination_types_covered"] = sorted(hal_types_covered)
    enh["n_benchmarks_mentioned"] = len(bm_mentioned)
    enh["n_benchmarks_with_scores"] = len(bm_scores)
    enh["analysis_timestamp"] = datetime.datetime.utcnow().isoformat() + "Z"
    enhanced.append(enh)

# ── 10. SAVE CSV ──────────────────────────────────────────────────────────────
print("Saving model_vs_benchmark_matrix.csv…")
bm_cols = [b for b in PRIMARY_BENCHMARKS if b in matrix_df.columns]
csv_out = matrix_df[["Organization","doc_type","modality"] + bm_cols].copy()
csv_out.to_csv(OUT / "model_vs_benchmark_matrix.csv", encoding="utf-8-sig")

# ── 11. HEATMAP ───────────────────────────────────────────────────────────────
print("Generating hallucination_type_heatmap.png…")

# Filter to models that have at least some coverage
haltype_plot = haltype_df.copy()
# Show only models with at least 1 hallucination type mentioned
active_models = haltype_plot[haltype_plot.sum(axis=1) > 0].index.tolist()
haltype_plot = haltype_plot.loc[active_models]

fig_h = max(12, len(haltype_plot) * 0.38)
fig, ax = plt.subplots(figsize=(16, fig_h))

cmap = matplotlib.colors.LinearSegmentedColormap.from_list(
    "hal", ["#f7f7f7", "#fddbc7", "#f4a582", "#d6604d", "#b2182b"])

sns.heatmap(
    haltype_plot,
    ax=ax,
    cmap=cmap,
    vmin=0, vmax=2,
    linewidths=0.4,
    linecolor="#dddddd",
    annot=True,
    fmt=".0f",
    annot_kws={"size": 7, "color": "#333333"},
    cbar_kws={"label": "Coverage (0=none, 1=mentioned, 2=with scores)"},
)
ax.set_title("VLM Hallucination Type Coverage Heatmap\n"
             "(based on official technical reports, system cards & model cards)",
             fontsize=14, pad=16, fontweight="bold")
ax.set_xlabel("Hallucination Type", fontsize=11)
ax.set_ylabel("Model", fontsize=11)
ax.xaxis.tick_top()
ax.xaxis.set_label_position("top")
plt.xticks(rotation=30, ha="left", fontsize=9)
plt.yticks(fontsize=8)
plt.tight_layout()
plt.savefig(OUT / "hallucination_type_heatmap.png", dpi=180, bbox_inches="tight")
plt.close()

# Also save a secondary heatmap: model × PRIMARY_BENCHMARKS (top 14)
print("Generating benchmark_coverage_heatmap.png…")
top14 = ["CHAIR","POPE","HallusionBench","MMHal-Bench","AMBER","FaithScore",
         "OCRBench","FACTS-Grounding","CharXiv","Video-MME",
         "LongVideoBench","TempCompass","SimpleVQA","OCRBench-v2"]
avail = [b for b in top14 if b in matrix_df.columns]
bm_heatmap = matrix_df[avail].copy()
active2 = bm_heatmap[bm_heatmap.sum(axis=1) > 0].index.tolist()
bm_heatmap = bm_heatmap.loc[active2]

fig2, ax2 = plt.subplots(figsize=(16, fig_h))
sns.heatmap(
    bm_heatmap,
    ax=ax2,
    cmap=cmap,
    vmin=0, vmax=2,
    linewidths=0.5,
    linecolor="#e0e0e0",
    annot=True,
    fmt=".0f",
    annot_kws={"size": 7, "color": "#333333"},
    cbar_kws={"label": "0=not reported, 1=mentioned, 2=scores reported"},
)
ax2.set_title("VLM × Hallucination Benchmark Coverage\n"
              "(primary hallucination / factuality benchmarks only)",
              fontsize=14, pad=16, fontweight="bold")
ax2.xaxis.tick_top()
ax2.xaxis.set_label_position("top")
plt.xticks(rotation=30, ha="left", fontsize=9)
plt.yticks(fontsize=8)
plt.tight_layout()
plt.savefig(OUT / "benchmark_coverage_heatmap.png", dpi=180, bbox_inches="tight")
plt.close()

# ── 12. EXCEL: official_experiment_results.xlsx ───────────────────────────────
print("Creating official_experiment_results.xlsx…")

def style_header(ws, row=1, fill="1F3864", font_color="FFFFFF"):
    fill_obj = PatternFill(fill_type="solid", fgColor=fill)
    font_obj = Font(bold=True, color=font_color, size=10)
    for cell in ws[row]:
        cell.fill = fill_obj
        cell.font = font_obj
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def auto_width(ws, min_w=8, max_w=40):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 2, max_w))

wb = Workbook()

# ── Sheet 1: Score Summary ──
ws1 = wb.active
ws1.title = "Score Summary"
ws1.freeze_panes = "F2"

# Headers
META_COLS = ["Organization","doc_type","source_file","modality"]
SCORE_COLS = list(SCORE_PATTERNS.keys())
all_cols = META_COLS + SCORE_COLS

ws1.append(["Model"] + all_cols)
style_header(ws1, row=1)

# Color scales for score cells
score_fill_colors = {
    (0, 60): "FFC7CE",   # red: low
    (60, 75): "FFEB9C",  # yellow: medium
    (75, 90): "C6EFCE",  # green: high
    (90, 101): "375623", # dark green: excellent
}

def score_fill(val):
    if val is None:
        return None
    for (lo, hi), color in score_fill_colors.items():
        if lo <= val < hi:
            return PatternFill(fill_type="solid", fgColor=color)
    return None

for model_name in results_df.index:
    row_data = results_df.loc[model_name]
    row = [model_name]
    for col in all_cols:
        if col in META_COLS:
            row.append(row_data.get(col, ""))
        else:
            v = row_data.get(col, None)
            row.append(v)
    ws1.append(row)

# Apply color fills to score cells
meta_offset = len(META_COLS) + 2  # model col + meta cols (1-indexed)
for row_idx, model_name in enumerate(results_df.index, start=2):
    for col_idx, col_name in enumerate(SCORE_COLS, start=meta_offset):
        val = results_df.loc[model_name, col_name] if col_name in results_df.columns else None
        cell = ws1.cell(row=row_idx, column=col_idx)
        if val is not None:
            fill = score_fill(val)
            if fill:
                cell.fill = fill

auto_width(ws1)
ws1.row_dimensions[1].height = 30

# ── Sheet 2: Full Matrix ──
ws2 = wb.create_sheet("Model×Benchmark Matrix")
ws2.freeze_panes = "E2"

full_cols = ["Organization","doc_type","modality"] + bm_cols
ws2.append(["Model"] + full_cols)
style_header(ws2, row=2, fill="1A3A5C")

for model_name in matrix_df.index:
    row_vals = [model_name]
    for col in full_cols:
        v = matrix_df.loc[model_name, col] if col in matrix_df.columns else 0
        row_vals.append(v)
    ws2.append(row_vals)

# Color cells by value (0=white, 1=yellow, 2=green)
color_map_matrix = {0: None, 1: "FFF2CC", 2: "E2EFDA"}
meta_end = 4  # Organization + doc_type + modality + Model = 4 cols
for row_idx in range(2, len(matrix_df) + 3):
    for col_idx in range(meta_end + 1, len(full_cols) + 2):
        cell = ws2.cell(row=row_idx, column=col_idx)
        try:
            v = int(cell.value or 0)
            if v in color_map_matrix and color_map_matrix[v]:
                cell.fill = PatternFill(fill_type="solid", fgColor=color_map_matrix[v])
        except Exception:
            pass

auto_width(ws2)

# ── Sheet 3: Hallucination Type Coverage ──
ws3 = wb.create_sheet("Hal Type Coverage")
ws3.append(["Model"] + HAL_TYPES)
style_header(ws3, row=1, fill="4F3A75")

cmap_hal = {0: None, 1: "FDD8B4", 2: "E05C2A"}
for model_name in haltype_df.index:
    row_vals = [model_name] + [haltype_df.loc[model_name, ht] for ht in HAL_TYPES]
    ws3.append(row_vals)
for row_idx in range(2, len(haltype_df) + 2):
    for col_idx in range(2, len(HAL_TYPES) + 2):
        cell = ws3.cell(row=row_idx, column=col_idx)
        try:
            v = int(cell.value or 0)
            if v and cmap_hal.get(v):
                cell.fill = PatternFill(fill_type="solid", fgColor=cmap_hal[v])
        except Exception:
            pass
auto_width(ws3)

# ── Sheet 4: Notes / Source Verification ──
ws4 = wb.create_sheet("Source Notes")
ws4.append(["Model","Organization","doc_type","source_type","source_url","local_path","modality","notes"])
style_header(ws4, row=1, fill="8B0000")

for entry in MODEL_ENTRIES:
    ws4.append([
        entry.get("model_or_benchmark_name",""),
        entry.get("organization",""),
        entry.get("doc_type",""),
        entry.get("source_type",""),
        entry.get("source_url",""),
        entry.get("local_path",""),
        entry.get("modality",""),
        entry.get("notes",""),
    ])
auto_width(ws4)

wb.save(OUT / "official_experiment_results.xlsx")
print("  Saved official_experiment_results.xlsx")

# ── 13. EXCEL: benchmark_taxonomy.xlsx ───────────────────────────────────────
print("Creating benchmark_taxonomy.xlsx…")
wb2 = Workbook()
ws_tax = wb2.active
ws_tax.title = "Benchmark Taxonomy"
ws_tax.freeze_panes = "B2"

tax_cols = ["full_name","arxiv","year","type","task","metrics",
            "scale","open_source","strengths","weaknesses",
            "common_pitfalls","hal_types","#Models Reporting"]
ws_tax.append(["Benchmark"] + tax_cols)
style_header(ws_tax, row=1, fill="1F3864")
ws_tax.row_dimensions[1].height = 28

type_colors = {
    "Object Hallucination": "FCE4D6",
    "Visual Hallucination": "FFF2CC",
    "OCR / Text Recognition Hallucination": "E2EFDA",
    "Video Understanding / Temporal Hallucination": "DAEEF3",
    "Chart Hallucination / Document Understanding": "EAD1DC",
    "Factual Grounding / Attribution Hallucination": "F2DCDB",
    "Faithfulness / Attribute Hallucination": "E6B8A2",
    "Comprehensive Hallucination": "CCC0DA",
    "Open-ended Multimodal Hallucination": "FFDAB9",
    "Temporal Comprehension Hallucination": "B0E0E6",
    "Long Video / Temporal Reasoning": "98FB98",
    "Object / Visual Hallucination": "FFFACD",
}

for bm_name, info in TAXONOMY.items():
    count_val = sum(1 for mn in ordered_models
                    if find_benchmark_mentions(model_texts.get(mn,""), bm_name))
    row_vals = [bm_name] + [info.get(c,"") for c in tax_cols[:-1]] + [count_val]
    ws_tax.append(row_vals)
    # Color by type
    bm_type = info.get("type","")
    color = type_colors.get(bm_type, "FFFFFF")
    for cell in ws_tax[ws_tax.max_row]:
        cell.fill = PatternFill(fill_type="solid", fgColor=color)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

auto_width(ws_tax, max_w=60)
ws_tax.column_dimensions["A"].width = 20
ws_tax.column_dimensions["B"].width = 40
ws_tax.column_dimensions["E"].width = 55
ws_tax.column_dimensions["F"].width = 35

wb2.save(OUT / "benchmark_taxonomy.xlsx")
print("  Saved benchmark_taxonomy.xlsx")

# ── 14. ENHANCED MANIFEST JSON ────────────────────────────────────────────────
print("Saving enhanced_manifest.json…")
(OUT / "enhanced_manifest.json").write_text(
    json.dumps(enhanced, ensure_ascii=False, indent=2))
print(f"  Saved {len(enhanced)} entries")

# ── 15. SUMMARY REPORT ───────────────────────────────────────────────────────
print("\n" + "="*60)
print("ANALYSIS COMPLETE — SUMMARY")
print("="*60)

hal_bms = ["CHAIR","POPE","HallusionBench","MMHal-Bench","AMBER",
           "FaithScore","OCRBench","Video-MME","CharXiv","FACTS-Grounding"]

print(f"\nModels analyzed: {len(ordered_models)}")
print(f"Benchmarks tracked: {len(BENCHMARKS)}")

print("\nTop models by hallucination benchmark coverage:")
bm_only = [b for b in hal_bms if b in matrix_df.columns]
coverage = matrix_df[bm_only].apply(lambda r: (r > 0).sum(), axis=1).sort_values(ascending=False)
for name, cnt in coverage.head(15).items():
    scores_cnt = (matrix_df.loc[name, bm_only] == 2).sum()
    print(f"  {name:35} {cnt}/{len(bm_only)} benchmarks  ({scores_cnt} with scores)")

print("\nHallucination benchmarks most commonly reported:")
for bm in hal_bms:
    if bm in matrix_df.columns:
        n_mention = (matrix_df[bm] >= 1).sum()
        n_scores  = (matrix_df[bm] == 2).sum()
        print(f"  {bm:25} mentioned={n_mention}  with_scores={n_scores}")

print(f"\nOutput files:")
for f in sorted(OUT.iterdir()):
    size_kb = f.stat().st_size // 1024
    print(f"  {f.name:50} {size_kb:>6} KB")
